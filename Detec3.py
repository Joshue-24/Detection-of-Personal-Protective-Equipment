import numpy as np
import matplotlib.pyplot as plt
from sklearn.model_selection import train_test_split
from sklearn.preprocessing import StandardScaler
from sklearn.linear_model import LogisticRegression
from sklearn.metrics import accuracy_score, confusion_matrix, classification_report
from ultralytics import YOLO
import cv2
import time
import csv
from datetime import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

# Cargar modelo de detección
model = YOLO("EPPS.pt")

# Cargar etiquetas
with open("etiquetas.txt", "r") as f:
    labels = [line.strip() for line in f.readlines()]

num_classes = len(model.model.names)
model.model.names = labels

# Inicializar cámara
cap = cv2.VideoCapture(2)

# Crear o abrir archivo XLSX
xlsx_file = "detecciones_epps.xlsx"
header = ["Fecha", "Hora"] + labels  # Encabezado del XLSX

try:
    workbook = load_workbook(xlsx_file)
    sheet = workbook.active
except FileNotFoundError:
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(header)  # Insertar encabezado si el archivo no existe

# Control de tiempo para guardar cada 10 min
start_time = time.time()

while True:
    ret, frame = cap.read()
    if not ret:
        print("Error: No se pudo leer el frame de la cámara")
        break

    # Realizar predicción con YOLO
    resultados = model.predict(frame, imgsz=416, conf=0.8)

    # Contar detecciones por tipo de EPP
    detecciones = {label: 0 for label in labels}
    for r in resultados:
        for box in r.boxes:
            cls_id = int(box.cls[0])  # Índice de la clase detectada
            class_name = labels[cls_id]
            detecciones[class_name] += 1

    # Mostrar anotaciones en la imagen
    anotaciones = resultados[0].plot()
    cv2.imshow("DETECCION EPPS", anotaciones)

    # Guardar en XLSX cada 10 minutos
    elapsed_time = time.time() - start_time
    if elapsed_time >= 5:  # 600 segundos = 10 minutos
        now = datetime.now()
        fecha = now.strftime("%Y-%m-%d")
        hora = now.strftime("%H:%M:%S")

        # Escribir datos en el XLSX
        row = [fecha, hora] + [detecciones[label] for label in labels]
        sheet.append(row)
        workbook.save(xlsx_file)
        
        print(f"[INFO] Datos guardados en {xlsx_file} a las {hora}")
        start_time = time.time()  # Reiniciar temporizador

    # Salir con la tecla ESC
    if cv2.waitKey(1) == 27:
        break

cap.release()
cv2.destroyAllWindows()
